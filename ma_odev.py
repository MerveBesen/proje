import sqlite3
from openpyxl import Workbook,load_workbook

ex_kurlar = load_workbook("C:/Users/Muhammet Sait/Desktop/kaynak/kurlar.xlsx")
kur = ex_kurlar.active

faizOranlari = load_workbook("C:/Users/Muhammet Sait/Desktop/kaynak/euro usd tl 6 12 12+ ay mevduat faizleri.xlsx")
faiz = faizOranlari.active

disBorc = load_workbook("C:/Users/Muhammet Sait/Desktop/kaynak/dış borç.xlsx")
borc = disBorc.active


enflasyon = load_workbook("C:/Users/Muhammet Sait/Desktop/kaynak/12-24 ay enflasyon beklentisi.xlsx")
enf = enflasyon.active

con = sqlite3.connect("C:/Users/Muhammet Sait/Desktop/Veriler.db")
cursor = con.cursor()



def add_kurlarTablosu():
    cursor.execute("CREATE TABLE IF NOT EXISTS Kurlar (Tarih TEXT, USD_Alış FLOAT, USD_Satış FLOAT, EUR_Alış FLOAT, EUR_Satış FLOAT, GBP_Alış FLOAT, GBP_Satış FLOAT)")
    con.commit()
def add_kurDegerleri(tarih,USD_Alış,USD_Satış,EUR_Alış,EUR_Satış,GBP_Alış,GBP_Satış):
    cursor.execute("INSERT INTO Kurlar VALUES (?,?,?,?,?,?,?)",(tarih,USD_Alış,USD_Satış,EUR_Alış,EUR_Satış,GBP_Alış,GBP_Satış))
    con.commit()


def add_faizOranlari():
    cursor.execute("CREATE TABLE IF NOT EXISTS FaizOranlari (Tarih TEXT, USD6_Faiz FLOAT, USD12_Faiz FLOAT, USDT_Faiz FLOAT, EUR6_Faiz FLOAT, EUR12_Faiz FLOAT, EURT_Faiz FLOAT, TL6_Faiz FLOAT, TL12_Faiz FLOAT, TLT_Faiz FLOAT)")
    con.commit()
def add_faizDegerleri(tarih,usd6,usd12,usdt,eur6,eur12,eurt,tl6,tl12,tlt):
    cursor.execute("INSERT INTO FaizOranlari VALUES (?,?,?,?,?,?,?,?,?,?)",(tarih,usd6,usd12,usdt,eur6,eur12,eurt,tl6,tl12,tlt))
    con.commit()


def add_disBorc():
    cursor.execute("CREATE TABLE IF NOT EXISTS Dis_Borc (Tarih TEXT, Borc FLOAT)")
    con.commit()
def add_borc(tarih,borc_):
    cursor.execute("INSERT INTO Dis_Borc VALUES (?,?)",(tarih,borc_))
    con.commit()
    

def addEnflasyon():
    cursor.execute("CREATE TABLE IF NOT EXISTS Enflasyon (Tarih TEXT, ENF12 FLOAT, ENF24 FLOAT)")
    con.commit()
def add_enfDeger(tarih,enf12,enf24):
    cursor.execute("INSERT INTO Enflasyon VALUES (?,?,?)",(tarih,enf12,enf24))
    con.commit()

add_kurlarTablosu()
for i in range(2,8387):
    add_kurDegerleri(tarih=kur.cell(i,1).value, USD_Alış=kur.cell(i,2).value, USD_Satış=kur.cell(i,3).value, EUR_Alış=kur.cell(i,4).value, EUR_Satış=kur.cell(i,5).value, GBP_Alış=kur.cell(i,6).value, GBP_Satış=kur.cell(i,7).value)

add_faizOranlari()
for i in range(2,1042):
    add_faizDegerleri(tarih=faiz.cell(i,1).value, usd6=faiz.cell(i,2).value, usd12=faiz.cell(i,3).value, usdt=faiz.cell(i,4).value, eur6=faiz.cell(i,5).value, eur12=faiz.cell(i,6).value, eurt=faiz.cell(i,7).value, tl6=faiz.cell(i,8).value, tl12=faiz.cell(i,9).value, tlt=faiz.cell(i,10).value)

add_disBorc()
for i in range(2,88):
    add_borc(tarih=borc.cell(i,1).value, borc_=borc.cell(i,2).value)

addEnflasyon()
for i in range(2,104):
    add_enfDeger(tarih=enf.cell(i,1).value, enf12=enf.cell(i,2).value, enf24=enf.cell(i,3).value)

con.close()

